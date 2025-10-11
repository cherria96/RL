import argparse
import os
import json
import pandas as pd
from openpyxl import load_workbook
from functions.Helper import data_interpolation, data_visualization, alphabet_to_number

# ----------------- args -----------------
parser = argparse.ArgumentParser(description="Process a BS workbook.")
parser.add_argument("FILE", help='Excel filename (e.g., "bsXXXX.xlsx"). '
                                 'If only a filename is given, it will be looked up under data/raw_data/.')
args = parser.parse_args()
FILE = args.FILE

# Normalize name and path
basename = os.path.basename(FILE)
name, ext = os.path.splitext(basename)
if not name.lower().startswith("bs"):
    raise ValueError(f'Filename must start with "bs": got "{name}"')

# If user supplied only a filename, prefix the standard folder
if os.path.isabs(FILE) or os.path.sep in FILE:
    file_path = FILE
else:
    file_path = os.path.join("data", "raw_data", FILE)

if not os.path.exists(file_path):
    raise FileNotFoundError(f'Input file not found: {file_path}')

# ----------------- config -----------------
with open("data/functions/config.json", "r") as f:
    config = json.load(f)

# Allow keys to be either the full path's basename or the exact FILE string
cfg_key = basename if basename in config else FILE
if cfg_key not in config:
    raise KeyError(f'Config for "{basename}" not found in data/functions/config.json')

# BS-specific config (we error out earlier if not bs*)
columns_to_remove = config[cfg_key]["columns_to_remove"]
use_cols = config[cfg_key]["use_cols"]
start, end = map(alphabet_to_number, use_cols.split(":"))
column_names = config[cfg_key]["column_names"]
drop_cols = config[cfg_key]["drop_cols"]
OUTPUT = config[cfg_key]["output"]

# ----------------- read sheets & headers -----------------
sheet_names = pd.ExcelFile(file_path).sheet_names
start_row = 14  # Row 15 in Excel (0-based index)

wb = load_workbook(file_path, data_only=True)
ws = wb[sheet_names[0]]
header_row1 = ws[15]  # Row 16 in Excel
header_row2 = ws[16]  # Row 17 in Excel

merged_headers = []
last_header = None
for cell1, cell2 in zip(header_row1[start:end+1], header_row2[start:end+1]):  # e.g., B to BH
    top = cell1.value
    sub = cell2.value

    if top is None:
        top = last_header
    else:
        last_header = top

    merged_header = str(top).strip() if top is not None else ""
    if sub is not None:
        merged_header = f"{merged_header}_{str(sub).strip()}"
    merged_headers.append(merged_header)

# ----------------- per-sheet processing -----------------
all_dfs = []
for sheet_name in sheet_names:
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=start_row, header=None, usecols=use_cols)
    df.columns = merged_headers

    # Optionally re-read two rows for checks (kept from your original code)
    df_check_row = pd.read_excel(file_path, sheet_name=sheet_name, header=None,
                                 skiprows=start_row, nrows=2, usecols=use_cols)
    second_header_row = df_check_row.iloc[1]  # not explicitly used, but kept

    # Filter out any columns with banned tokens
    cols_to_keep = [col for col in merged_headers if not any(ex in col for ex in columns_to_remove)]
    df = df[cols_to_keep]

    # Map to canonical column names then drop unwanted columns
    df.columns = column_names
    df = df.drop(columns=drop_cols, errors="ignore")

    all_dfs.append(df)

# ----------------- concatenate & clean -----------------
concatenated_df = pd.concat(all_dfs, axis=0, ignore_index=True)
filtered_df = concatenated_df.dropna(subset=['Date'])
filtered_df['Date'] = pd.to_datetime(filtered_df['Date'], format='%Y-%m-%d', errors='coerce')
filtered_df.dropna(subset=['Date'], inplace=True)
filtered_df.iloc[:, 1:] = filtered_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')

# ----------------- write -----------------
out_path = os.path.join("data", "processed_data", OUTPUT)
os.makedirs(os.path.dirname(out_path), exist_ok=True)
filtered_df.to_csv(out_path, index=False)
print(f"Saved: {out_path}")
